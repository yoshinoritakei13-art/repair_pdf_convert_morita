"""
修理見積PDF → Excel 変換ツール
メーカー（モリタ）からのPDFを読み込み、必要項目をExcelに書き出す
"""

import os
import re
import io
from datetime import datetime

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _normalize_text(text: str) -> str:
    """
    PDF抽出文字の全角英数字を半角に寄せて、正規表現のマッチを安定させる。
    """
    zenkaku = "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
    hankaku = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
    text_n = text.translate(str.maketrans(zenkaku, hankaku))
    return text_n.upper()


def _extract_request_no(page, text_n: str) -> str:
    """
    依頼番号（英字2文字+英数字10文字=計12文字）を「依頼内容」右側の同一行から拾う。
    pdfplumberの座標情報を使って、WD等の別欄を拾う誤爆を避ける。
    """
    pat = re.compile(r"^[A-Z]{2}[0-9A-Z]{10}$")

    # まずはwords抽出（同一行の右側を拾いやすい）
    try:
        words = page.extract_words(use_text_flow=True)
    except Exception:
        words = []

    # 「依頼内容」文字を探す（スペース混入に強くする）
    target = None
    for w in words:
        t = _normalize_text(w.get("text", ""))
        t_compact = re.sub(r"\s+", "", t)
        if "依頼内容" in t_compact:
            target = w
            break

    if target:
        top = float(target.get("top", 0.0))
        bottom = float(target.get("bottom", top))
        x1 = float(target.get("x1", 0.0))

        # 同じ行（topが近い）かつ右側のワードを集めて、依頼番号候補を探す
        right_words = []
        for w in words:
            if w.get("x0", 0.0) <= x1:
                continue
            if abs(w.get("top", 0.0) - top) > 3.0:
                continue
            right_words.append(w)

        # 右側ワードを左→右で連結し、12文字コードをスペース無視で探索
        right_words.sort(key=lambda x: x.get("x0", 0.0))
        joined = " ".join(w.get("text", "") for w in right_words)
        joined_n = _normalize_text(joined)
        # スペース混入に対応して、文字間スペースを許容しつつ12文字を復元
        m = re.search(r"([A-Z]{2}(?:\s*[0-9A-Z]){10})", joined_n)
        if m:
            cand = re.sub(r"\s+", "", m.group(1))
            if pat.match(cand):
                return cand

        # ワード単体でも12文字候補があれば採用
        for w in right_words:
            cand = re.sub(r"\s+", "", _normalize_text(w.get("text", "")))
            if pat.match(cand):
                return cand

        # 右側の“欄”が別行扱いになるPDF向けに、依頼内容の右側矩形を切り出して抽出
        try:
            page_w = float(getattr(page, "width", 0.0)) or 0.0
            y0 = max(0.0, top - 12.0)
            y1 = bottom + 18.0
            x0 = max(0.0, x1 + 2.0)
            x1r = page_w
            if page_w > 0 and x0 < x1r:
                cropped = page.crop((x0, y0, x1r, y1))
                cropped_text = cropped.extract_text() or ""
                cropped_n = _normalize_text(cropped_text)
                m3 = re.search(r"([A-Z]{2}(?:\s*[0-9A-Z]){10})", cropped_n)
                if m3:
                    cand = re.sub(r"\s+", "", m3.group(1))
                    if pat.match(cand) and not cand.startswith("WD"):
                        return cand
        except Exception:
            pass

    # フォールバック: テキストから「依 頼 内 容」などの分割も許容して近傍を抽出
    anchor = re.search(r"依\s*頼\s*内\s*容", text_n)
    if anchor:
        window = text_n[anchor.start() : min(len(text_n), anchor.start() + 500)]
        m = re.search(r"([A-Z]{2}(?:\s*[0-9A-Z]){10})", window)
        if m:
            cand = re.sub(r"\s+", "", m.group(1))
            if pat.match(cand) and not cand.startswith("WD"):
                return cand

    return ""


def parse_dealer_page(text: str):
    """ディーラー様用ページから必要項目を抽出する（ユーザー様用は無視）"""
    if "ディーラー様用" not in text:
        return None

    text_n = _normalize_text(text)
    lines = text_n.split("\n")

    data: dict = {}

    # 修理受付番号
    receipt_match = re.search(r"修理受付[Nn][Oo]\.?\s*([A-Z]{2}\d+)", text_n)
    data["receipt_no"] = receipt_match.group(1) if receipt_match else ""

    # 歯科医院/技工所名
    clinic_match = re.search(r"歯科医院／技工所\s+(.+?)\s+様", text_n)
    if not clinic_match:
        return None
    data["clinic"] = clinic_match.group(1).strip()

    # 製造番号
    mfg_match = re.search(r"製造番号\s+(\S+)", text_n)
    data["mfg_no"] = mfg_match.group(1).strip() if mfg_match else ""

    # 依頼番号はページ座標からの抽出に移行（parse_dealer_pageでは埋めない）
    data["request_no"] = ""

    # 小計（卸）: 「小　計 〇〇」の最後の数値（必須小計・推奨小計を除く意図）
    subtotal_matches = re.findall(r"小\s*計\s+([\d,]+)", text_n)
    subtotal_value = int(subtotal_matches[-1].replace(",", "")) if subtotal_matches else 0
    data["subtotal"] = subtotal_value

    # No列（1,2,3,...）の行から卸価を拾う。
    # 折り返しで金額が次行に出ることがあるため、「No行 + 次行以降（次のNoが出るまで）」を連結して解析する。
    items: dict[int, int] = {}
    expected_no = 1
    i = 0
    started = False

    def starts_with_no(s: str) -> int | None:
        m = re.match(r"^(\d{1,2})\b", s)
        if not m:
            return None
        try:
            n = int(m.group(1))
        except ValueError:
            return None
        return n if 1 <= n <= 20 else None

    while i < len(lines):
        s = lines[i].strip()
        if not s:
            if started:
                i += 1
                continue
            i += 1
            continue

        no = starts_with_no(s)
        if no is None:
            i += 1
            continue

        if not started:
            if no != 1:
                i += 1
                continue
            started = True
            expected_no = 1

        if no != expected_no:
            # Noが連番で途切れた＝空欄になった扱いで終了
            break

        # ブロックを「次のNo行が出る直前」まで連結
        block = s
        j = i + 1
        while j < len(lines):
            t = lines[j].strip()
            if not t:
                # 空行はブロック継続（折り返し対策）
                j += 1
                continue
            # 表の終了（小計/合計など）に入ったら、ここでブロックを止める
            if re.search(r"(小\s*計|合\s*計|総\s*計)", t):
                break
            n2 = starts_with_no(t)
            if n2 is not None:
                break
            block += " " + t
            j += 1

        numbers = re.findall(r"\d[\d,]*", block)
        nums = [int(x.replace(",", "")) for x in numbers]
        # 数量(1)などの小さい数値を除外して金額候補だけに寄せる。
        # また、小計/合計の値を誤って拾わないように除外する。
        price_candidates = [n for n in nums if n != no and n >= 100 and n != subtotal_value]
        if price_candidates:
            # 末尾2つが「卸価」「標準価格」になりやすいので、2つあるなら小さい方を卸価
            wholesale = min(price_candidates[-2], price_candidates[-1]) if len(price_candidates) >= 2 else price_candidates[-1]
            items[no] = wholesale

        expected_no += 1
        i = j

    data["items"] = items

    return data


def create_excel(results: list[dict], output_path: str):
    """抽出データをExcelファイルに書き出す"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "修理見積"

    # 最大NO番号を確認
    max_no = 0
    for r in results:
        if r.get("items"):
            max_no = max(max_no, max(r["items"].keys()))
    max_no = max(max_no, 5)  # 最低5列確保

    # ヘッダー作成
    headers = ["修理受付番号", "歯科医院/技工所名", "製造番号", "依頼番号"]
    for i in range(1, max_no + 1):
        headers.append(f"NO{i}卸価格")
    headers.append("小計(卸)")

    # ヘッダースタイル
    header_fill = PatternFill(start_color="17375E", end_color="17375E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 30

    # データ行のスタイル
    data_font = Font(size=10)
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    num_format = "#,##0"

    # 交互行の背景色
    fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_odd = PatternFill(fill_type=None)

    for row_idx, data in enumerate(results, 2):
        row_fill = fill_even if row_idx % 2 == 0 else fill_odd

        def set_cell(col: int, value, align=left_align, fmt: str | None = None):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = align
            cell.border = border
            cell.fill = row_fill
            if fmt:
                cell.number_format = fmt

        set_cell(1, data.get("receipt_no", ""))
        set_cell(2, data.get("clinic", ""))
        set_cell(3, data.get("mfg_no", ""))
        set_cell(4, data.get("request_no", ""))

        items = data.get("items", {})
        for no in range(1, max_no + 1):
            col = 4 + no
            if no in items:
                set_cell(col, items[no], right_align, num_format)
            else:
                set_cell(col, None, right_align)

        set_cell(4 + max_no + 1, data.get("subtotal") or None, right_align, num_format)

    # 列幅調整
    col_widths = {"A": 15, "B": 35, "C": 16, "D": 18}
    for k, v in col_widths.items():
        ws.column_dimensions[k].width = v
    for i in range(5, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path


def create_excel_bytes(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "修理見積"

    max_no = 0
    for r in results:
        if r.get("items"):
            max_no = max(max_no, max(r["items"].keys()))
    max_no = max(max_no, 5)

    headers = ["修理受付番号", "歯科医院/技工所名", "製造番号", "依頼番号"]
    for i in range(1, max_no + 1):
        headers.append(f"NO{i}卸価格")
    headers.append("小計(卸)")

    header_fill = PatternFill(start_color="17375E", end_color="17375E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 30

    data_font = Font(size=10)
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    num_format = "#,##0"

    fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_odd = PatternFill(fill_type=None)

    for row_idx, data in enumerate(results, 2):
        row_fill = fill_even if row_idx % 2 == 0 else fill_odd

        def set_cell(col: int, value, align=left_align, fmt: str | None = None):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = align
            cell.border = border
            cell.fill = row_fill
            if fmt:
                cell.number_format = fmt

        set_cell(1, data.get("receipt_no", ""))
        set_cell(2, data.get("clinic", ""))
        set_cell(3, data.get("mfg_no", ""))
        set_cell(4, data.get("request_no", ""))

        items = data.get("items", {})
        for no in range(1, max_no + 1):
            col = 4 + no
            if no in items:
                set_cell(col, items[no], right_align, num_format)
            else:
                set_cell(col, None, right_align)
        set_cell(4 + max_no + 1, data.get("subtotal") or None, right_align, num_format)

    col_widths = {"A": 15, "B": 35, "C": 16, "D": 18}
    for k, v in col_widths.items():
        ws.column_dimensions[k].width = v
    for i in range(5, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def convert_pdf_to_results(pdf_file) -> tuple[list[dict], int]:
    results: list[dict] = []
    with pdfplumber.open(pdf_file) as pdf:
        total_pages = len(pdf.pages)
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                data = parse_dealer_page(text)
                if data:
                    data["request_no"] = _extract_request_no(page, _normalize_text(text))
                    results.append(data)
    return results, total_pages


def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    pdf_path = filedialog.askopenfilename(
        title="修理見積PDFを選択してください",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
    )

    if not pdf_path:
        messagebox.showinfo("キャンセル", "ファイルが選択されませんでした。")
        root.destroy()
        return

    try:
        results: list[dict] = []
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    data = parse_dealer_page(text)
                    if data:
                        # 依頼番号はページ座標から確実に抽出して上書き
                        data["request_no"] = _extract_request_no(page, _normalize_text(text))
                        results.append(data)

        if not results:
            messagebox.showwarning(
                "データなし",
                "ディーラー様用ページが見つかりませんでした。\nPDFの内容を確認してください。",
            )
            root.destroy()
            return

        # 出力ファイル名
        today = datetime.now().strftime("%Y%m%d")
        pdf_dir = os.path.dirname(pdf_path)
        output_path = os.path.join(pdf_dir, f"修理見積_{today}.xlsx")

        # 同名ファイルが存在する場合は連番
        counter = 1
        while os.path.exists(output_path):
            output_path = os.path.join(pdf_dir, f"修理見積_{today}_{counter}.xlsx")
            counter += 1

        create_excel(results, output_path)

        messagebox.showinfo(
            "変換完了",
            "Excelファイルを作成しました！\n\n"
            f"　総ページ数: {total_pages}ページ\n"
            f"　抽出件数: {len(results)}件\n\n"
            f"保存先:\n{output_path}",
        )

        if hasattr(os, "startfile"):
            os.startfile(output_path)

    except Exception as e:
        import traceback

        messagebox.showerror(
            "エラー",
            f"エラーが発生しました:\n{str(e)}\n\n詳細:\n{traceback.format_exc()}",
        )

    root.destroy()


if __name__ == "__main__":
    main()

